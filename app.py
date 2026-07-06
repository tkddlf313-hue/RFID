"""
============================================================
 스마트 팩토리 UHF RFID 재고조사 시스템  v2.1
 흐름: 입고·태그발행 → 출고스캔(QR) → 재고조사 스캔 → 엑셀 출력
============================================================
"""

# ── 표준 라이브러리 ────────────────────────────────────────
import base64
import datetime
import io
import json
import os
import re
import socket
import sqlite3
import threading
import time
import uuid

# ── 서드파티 ───────────────────────────────────────────────
import cv2
import numpy as np
import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter
import pandas as pd
import serial
import serial.tools.list_ports
import streamlit as st
from google import genai
from google.genai import types
try:
    import qrcode
    from PIL import Image as _PIL_Image, ImageDraw, ImageFont
    _QR_AVAILABLE = True
except ImportError:
    _QR_AVAILABLE = False



# ══════════════════════════════════════════════════════════
# 상수 및 경로 설정
# ══════════════════════════════════════════════════════════
# 실행 경로에 구애받지 않도록 파일 기준 절대경로 계산
BASE_DIR        = os.path.dirname(os.path.abspath(__file__))
DB_PATH         = os.path.join(BASE_DIR, "rfid_inventory.db")
MODELS_DIR      = os.path.join(BASE_DIR, "models")  # 폐쇄망 대비 로컬 모델 폴더
YOLO_CANDIDATES = [                   # 우선순위 순서로 시도
    "yolo11n.pt",
    "yolo11s.pt",
    "yolo11m.pt",
]
# 화면 갱신: 리더기 연결 중 polling 간격(초) — sleep 없이 st_autorefresh 대체
_POLL_SEC = 2


# ══════════════════════════════════════════════════════════
# DB 초기화
# ══════════════════════════════════════════════════════════
def get_db() -> sqlite3.Connection:
    """매 호출마다 새 연결 반환 (멀티스레드 안전)"""
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn


# 자재 분류 목록
CATEGORY_LIST = [
    "원재료",
    "부재료(조성)",
    "부재료(가공)",
    "부재료(환경)",
    "기자재(초지)",
    "기자재(가공)",
    "기자재(완정)",
    "기자재(보전)",
    "기자재(안전)",
]

# 더미 데이터 (앱 최초 실행 시 DB가 비어있으면 자동 삽입)
# (epc, 자재명, 자재번호, 배치코드, 수량, 분류, 저장위치)
_DUMMY_TAGS = [
    ("E004015025A1C00100000001", "FELT:3.84M*23.40M,PM52 1P", "603160", "10106", 4, "기자재(초지)", "8300"),
    ("E004015025A1C00200000002", "FELT:3.80M*17.50M,PM52 2P", "603161", "10106", 5, "기자재(초지)", "8300"),
    ("E004015025A1C00300000003", "FELT:3.80M*16.00M,PM52 3P", "603162", "10106", 4, "기자재(초지)", "8300"),
    ("E004015025A1C00400000004", "CANVAS:3.75M*31.00M,PM52 1,2군 SINGLE", "603168", "10108", 1, "기자재(초지)", "8300"),
    ("E004015025A1C00500000005", "CANVAS:3.75M*31.00M,PM52 1,2군 SINGLE", "603168", "10114", 1, "기자재(초지)", "8300"),
    ("E004015025A1C00600000006", "CANVAS:3.75M*31.00M,PM52 5군 TOP", "603169", "10101", 2, "기자재(초지)", "8300"),
    ("E004015025A1C00700000007", "CANVAS:3.75M*31.00M,PM52 5군 TOP", "603169", "10114", 1, "기자재(초지)", "8300"),
    ("E004015025A1C00800000008", "CANVAS:3.75M*38.00M,PM52 3,4군 BTM", "603171", "10114", 3, "기자재(초지)", "8300"),
    ("E004015025A1C00900000009", "ROD:GROOVED,Ø25*15*4120,RAUA504987", "605163", "", 18, "기자재(초지)", "8300"),
    ("E004015025A1C01000000010", "ROD:GROOVED,Ø25*23*4120,RAUA504988", "605164", "", 74, "기자재(초지)", "8300"),
    ("E004015025A1C01100000011", "ROD:GROOVED,VMRØ25*13*4120,RAUA505123", "605254", "", 19, "기자재(초지)", "8300"),
    ("E004015025A1C01200000012", "ROD BAR:RAUA505124,APPLICATORØ25*4120", "605258", "", 26, "기자재(초지)", "8300"),
    ("E004015025A1C01300000013", "ROD BED:Ø25,RAUA504989", "605303", "", 27, "기자재(초지)", "8300"),
    ("E004015025A1C01400000014", "ROD:GROOVED,Ø25*30*4120,RAUA506044", "616785", "", 14, "기자재(초지)", "8300"),
    ("E004015025A1C01500000015", "CANVAS:3.75M*34.00M,PM52,5군,BOTTOM", "618004", "10101", 2, "기자재(초지)", "8300"),
    ("E004015025A1C01600000016", "ROD:GROOVED,Ø25*35*4120,RAUA506576", "618278", "", 29, "기자재(초지)", "8300"),
    ("E004015025A1C01700000017", "WIRE:3.87M*35.25M,PM52 3.0L", "618771", "10102", 3, "기자재(초지)", "8300"),
    ("E004015025A1C01800000018", "VANE:450*3707(PVC),RAUA122137", "618949", "", 3, "기자재(보전)", "8300"),
    ("E004015025A1C01900000019", "CANVAS:3.77M*30.50M, PM52, 1군 SINGLE", "619152", "10112", 2, "기자재(초지)", "8300"),
    ("E004015025A1C02000000020", "CANVAS:3.75M*30.50M, PM52, 2군 SINGLE", "619153", "10112", 2, "기자재(초지)", "8300"),
    ("E004015025A1C02100000021", "CANVAS:3.75M*34.50M, PM52, 3/4군 TOP", "619154", "10101", 1, "기자재(초지)", "8300"),
    ("E004015025A1C02200000022", "CANVAS:3.75M*34.50M, PM52, 3/4군 TOP", "619154", "10114", 2, "기자재(초지)", "8300"),
    ("E004015025A1C02300000023", "HOSE,LOADING:CR L=9000", "626348", "", 2, "기자재(보전)", "8300"),
    ("E004015025A1C02400000024", "TUBE,AIR:25*30M,고온용,GL&V", "626378", "", 1, "기자재(보전)", "8300"),
    ("E004015025A1C02500000025", "EXPANDER T BAR:Ø20*3540L,SUJ2,PM52", "626380", "", 1, "기자재(완정)", "8300"),
]


def init_db() -> None:
    """앱 시작 시 테이블이 없으면 생성하고, 기존 DB 마이그레이션 및 더미 데이터 삽입"""
    with get_db() as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS tags (
                tag_id      TEXT PRIMARY KEY,
                item_name   TEXT NOT NULL,
                lot_number  TEXT NOT NULL,
                quantity    INTEGER DEFAULT 0,
                issued_at   TEXT NOT NULL,
                location    TEXT DEFAULT '',
                category    TEXT DEFAULT '미분류',
                mat_number  TEXT DEFAULT '',
                batch_code  TEXT DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS scan_log (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                tag_id      TEXT NOT NULL,
                scanned_at  TEXT NOT NULL,
                matched     INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS outgoing_log (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                tag_id     TEXT NOT NULL,
                item_name  TEXT NOT NULL,
                category   TEXT DEFAULT '',
                mat_number TEXT DEFAULT '',
                qty_out    INTEGER NOT NULL,
                qty_before INTEGER NOT NULL,
                qty_after  INTEGER NOT NULL,
                out_at     TEXT NOT NULL,
                reason     TEXT DEFAULT '소진'
            );
            CREATE TABLE IF NOT EXISTS incoming_log (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                tag_id     TEXT NOT NULL,
                item_name  TEXT NOT NULL,
                category   TEXT DEFAULT '',
                mat_number TEXT DEFAULT '',
                qty_in     INTEGER NOT NULL,
                qty_before INTEGER NOT NULL,
                qty_after  INTEGER NOT NULL,
                in_at      TEXT NOT NULL,
                remark     TEXT DEFAULT '입고'
            );
        """)
        # ── 기존 DB 마이그레이션: 없는 컬럼 추가 ──
        cols = [r[1] for r in conn.execute("PRAGMA table_info(tags)").fetchall()]
        if "category" not in cols:
            conn.execute("ALTER TABLE tags ADD COLUMN category TEXT DEFAULT '미분류'")
        if "mat_number" not in cols:
            conn.execute("ALTER TABLE tags ADD COLUMN mat_number TEXT DEFAULT ''")
        if "batch_code" not in cols:
            conn.execute("ALTER TABLE tags ADD COLUMN batch_code TEXT DEFAULT ''")
        # ── DB가 비어있으면 더미 데이터 자동 삽입 ──
        if conn.execute("SELECT COUNT(*) FROM tags").fetchone()[0] == 0:
            import datetime as _dt, random as _rnd
            for epc, name, mat_no, batch, qty, cat, loc in _DUMMY_TAGS:
                delta = _dt.timedelta(
                    days=_rnd.randint(0, 30),
                    hours=_rnd.randint(8, 17),
                    minutes=_rnd.randint(0, 59),
                )
                issued = (_dt.datetime.now() - delta).strftime("%Y-%m-%d %H:%M:%S")
                conn.execute(
                    "INSERT OR IGNORE INTO tags(tag_id,item_name,lot_number,quantity,issued_at,location,category,mat_number,batch_code) VALUES (?,?,?,?,?,?,?,?,?)",
                    (epc, name, batch, qty, issued, loc, cat, mat_no, batch),
                )


def db_reset_dummy() -> None:
    """기존 데이터를 모두 삭제하고 신규 더미 데이터로 초기화"""
    with get_db() as conn:
        conn.execute("DELETE FROM tags")
        conn.execute("DELETE FROM scan_log")
        conn.execute("DELETE FROM outgoing_log")
        conn.execute("DELETE FROM incoming_log")
        import datetime as _dt, random as _rnd
        for epc, name, mat_no, batch, qty, cat, loc in _DUMMY_TAGS:
            delta = _dt.timedelta(
                days=_rnd.randint(0, 30),
                hours=_rnd.randint(8, 17),
                minutes=_rnd.randint(0, 59),
            )
            issued = (_dt.datetime.now() - delta).strftime("%Y-%m-%d %H:%M:%S")
            conn.execute(
                "INSERT OR IGNORE INTO tags(tag_id,item_name,lot_number,quantity,issued_at,location,category,mat_number,batch_code) VALUES (?,?,?,?,?,?,?,?,?)",
                (epc, name, batch, qty, issued, loc, cat, mat_no, batch),
            )


init_db()


# ══════════════════════════════════════════════════════════
# 페이지 기본 설정
# ══════════════════════════════════════════════════════════
st.set_page_config(
    page_title="RFID 재고조사 시스템",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded",
)

# 상단 헤더
st.markdown(
    """
    <div style='background:linear-gradient(90deg,#1B365D,#2E5FA3);
                padding:1.2rem 1.5rem; border-radius:10px; margin-bottom:1rem;'>
        <h2 style='color:#FFFFFF; margin:0;'>📦 스마트 팩토리 UHF RFID 재고조사 시스템</h2>
        <h5 style='color:#E6EEF8; margin:0.3rem 0 0; font-weight:normal;'>RFID 기반 상시 재고조사 체계 구축을 통한 천안공장 창고 효율화 전략</h5>
        <p style='color:#B0C4DE; margin:0.4rem 0 0; font-size:13.5px; font-weight:bold;'>
            🏆 한솔 AI Festival &nbsp;|&nbsp; 한솔제지 Team 천안YB (황원민, 한상일, 박준성)
        </p>
    </div>
    """,
    unsafe_allow_html=True,
)

# 맥킨지 스타일의 요약 박스 (Executive Summary)
st.markdown(
    """
    <div style='background-color:#F0F4F8; border-left:5px solid #1B365D; padding:12px 18px; border-radius:4px; margin-bottom:1.5rem;'>
        <div style='display:flex; justify-content:space-between; margin-bottom:8px;'>
            <span style='font-size:11px; font-weight:bold; color:#1B365D; text-transform:uppercase; letter-spacing:1px;'>Executive Summary</span>
            <span style='font-size:11px; color:#555; font-weight:555;'>한솔제지 천안공장</span>
        </div>
        <div style='font-size:14px; font-weight:bold; color:#111; margin-bottom:6px;'>UHF RFID와 QR코드를 결합한 하이브리드 스마트 창고 관리 플랫폼</div>
        <div style='font-size:12.5px; color:#333; line-height:1.6;'>
            • <b>현장의 문제 해결</b>: 원부재료, 기자재의 수작업 재고조사 한계 및 선입선출 추적 오류 해결<br/>
            • <b>하이브리드 프로세스</b>: <u>입고/재고조사</u>는 대용량 무선 RFID로, <u>입/출고 기록</u>은 스마트폰 카메라 QR 스캔으로 처리하여 비용 및 작업 편의성 최적화<br/>
            • <b>기대 효과</b>:<br/>
            &nbsp;&nbsp;1. 창고 실시간 전수 재고조사 시간 90% 이상 단축 및 선입선출 오류 제로화 달성<br/>
            &nbsp;&nbsp;2. 고가 부재료(Felt, Canvas 등)의 선입선출(FIFO) 엄수를 통한 자재 사용 수명 극대화 및 안전 재고 최적화로 창고 운전자본 효율성 대폭 증가
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)


# ══════════════════════════════════════════════════════════
# 세션 상태 초기화
# ══════════════════════════════════════════════════════════
_SESSION_DEFAULTS: dict = {
    "connected":    False,
    "conn_type":    "시리얼 (COM)",
    "serial_obj":   None,
    "tcp_socket":   None,
    "stop_flag":    False,
    "log_messages": [],
    "last_rfid":    None,
    "scan_session": [],       # 현재 재고조사 세션 스캔 결과 누적
    "scan_count_prev": 0,     # 이전 렌더링 시 스캔 건수 (신규 감지 토스트용)
}
for _k, _v in _SESSION_DEFAULTS.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v


# ══════════════════════════════════════════════════════════
# 공통 유틸
# ══════════════════════════════════════════════════════════
def log(msg: str) -> None:
    """시스템 로그 추가 (최대 100줄 유지)"""
    ts = datetime.datetime.now().strftime("%H:%M:%S")
    st.session_state["log_messages"].append(f"[{ts}] {msg}")
    st.session_state["log_messages"] = st.session_state["log_messages"][-100:]


def gen_epc() -> str:
    """UUID 기반 96-bit EPC(24자리 HEX) 자동 생성"""
    return uuid.uuid4().hex[:24].upper()


def now_str() -> str:
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def extract_qty_from_texts(texts: list[str]) -> int | None:
    """
    EasyOCR 텍스트 리스트에서 수량 숫자 추출.
    '수량', 'QTY', 'EA', 'PCS' 키워드 우선, 없으면 최댓값 반환.
    """
    full = " ".join(texts)
    for pat in [
        r'(?:수량|QTY|qty)[:\s]*(\d+)',
        r'(\d+)\s*(?:EA|PCS|개|ea|pcs)',
    ]:
        m = re.search(pat, full)
        if m:
            return int(m.group(1))
    nums = re.findall(r'\b(\d{1,6})\b', full)
    return max(int(n) for n in nums) if nums else None


# ══════════════════════════════════════════════════════════
# Gemini 2.5 Flash OCR
# ══════════════════════════════════════════════════════════
def get_gemini_keys(user_key: str) -> list[str]:
    """사용자 입력 키와 secrets.toml에 등록된 키들을 조합하여 리스트로 반환"""
    keys = []
    if user_key and "YOUR_GEMINI" not in user_key:
        # 사용자가 직접 세미콜론, 쉼표, 줄바꿈으로 여러 개를 입력했을 수 있으므로 분할 지원
        keys.extend([k.strip() for k in re.split(r'[,;\s\n\r]+', user_key) if k.strip()])
    
    # secrets.toml에서 가져오기
    if "GEMINI_API_KEYS" in st.secrets:
        toml_keys = st.secrets["GEMINI_API_KEYS"]
        if isinstance(toml_keys, list):
            for tk in toml_keys:
                if tk and "YOUR_GEMINI" not in tk and tk not in keys:
                    keys.append(tk)
        elif isinstance(toml_keys, str) and toml_keys not in keys:
            keys.append(toml_keys)
    
    if "GEMINI_API_KEY" in st.secrets:
        tk = st.secrets["GEMINI_API_KEY"]
        if tk and "YOUR_GEMINI" not in tk and tk not in keys:
            keys.append(tk)
            
    return [k for k in keys if k]


# ══════════════════════════════════════════════════════════
# Gemini 2.5 Flash OCR
# ══════════════════════════════════════════════════════════
def gemini_ocr(image_bytes: bytes, api_keys: list[str] | str) -> dict:
    """
    Gemini 2.5 Flash Vision으로 자재 라벨 이미지를 분석하여
    품명·Lot번호·수량을 JSON으로 반환.
    API 키 리스트를 돌며 하나가 한도 초과 등으로 실패하면 다음 키로 자동 재시도합니다.
    """
    keys = [api_keys] if isinstance(api_keys, str) else api_keys
    if not keys:
        return {"raw_text": "오류: 활성화된 Gemini API 키가 없습니다.", "item_name": None, "lot": None, "quantity": None}

    last_err = None
    for idx, key in enumerate(keys):
        try:
            client = genai.Client(api_key=key)
            prompt = """이 이미지는 공장 자재 라벨입니다.
아래 JSON 형식으로만 응답해 주세요. 값을 찾을 수 없으면 null로 표기하세요.
{
  "raw_text":  "라벨 전체 텍스트",
  "item_name": "품명",
  "lot":       "Lot 번호",
  "quantity":  숫자만 (단위 제외, 정수)
}
수량은 '수량', 'QTY', 'EA', 'PCS', '개' 키워드 근처 숫자를 우선합니다."""

            response = client.models.generate_content(
                model="gemini-2.5-flash",
                contents=[
                    types.Part.from_bytes(data=image_bytes, mime_type="image/jpeg"),
                    prompt,
                ],
            )

            raw_text = response.text.strip()
            # 응답에서 JSON 블록만 추출
            m = re.search(r'\{.*\}', raw_text, re.DOTALL)
            if m:
                try:
                    return json.loads(m.group())
                except json.JSONDecodeError:
                    pass
            # JSON 파싱 실패 시 raw_text만 담아 반환
            return {"raw_text": raw_text, "item_name": None, "lot": None, "quantity": None}
        except Exception as e:
            last_err = e
            log(f"[Gemini OCR] {idx+1}번째 API 키 실패 (다음 키 시도): {e}")
            continue

    # 모든 키가 실패했을 경우 예외 발생
    raise Exception(f"모든 Gemini API 키가 실패했습니다. 마지막 오류: {last_err}")


def gemini_count_objects(image_bytes: bytes, target_object: str, api_keys: list[str] | str) -> dict:
    """
    Gemini 2.5 Flash를 사용하여 이미지 내 특정 물체의 개수를 카운팅하고
    바운딩 박스 좌표 정보를 포함하여 결과를 반환합니다.
    API 키 리스트를 돌며 하나가 한도 초과 등으로 실패하면 다음 키로 자동 재시도합니다.
    """
    keys = [api_keys] if isinstance(api_keys, str) else api_keys
    if not keys:
        return {"target_object": target_object, "total_count": 0, "detected_objects": []}

    last_err = None
    for idx, key in enumerate(keys):
        try:
            client = genai.Client(api_key=key)

            prompt = f"""이 이미지는 공장 자재 또는 물품 보관 이미지입니다.
이미지에서 '{target_object}'(또는 이에 해당하는 사물)의 개수를 세어주세요.
그리고 감지된 각 '{target_object}'의 위치(바운딩 박스) 정보를 알려주세요.
반드시 아래 JSON 형식으로만 응답해 주세요. 마크다운 ```json ``` 블록 없이 순수 JSON만 응답해 주세요.

{{
  "target_object": "{target_object}",
  "total_count": 물체 개수 (숫자),
  "detected_objects": [
    {{
      "box_2d": [ymin, xmin, ymax, xmax],  # 0~1000 사이의 상대 좌표값 (정수)
      "label": "물체 라벨"
    }}
  ]
}}
"""

            response = client.models.generate_content(
                model="gemini-2.5-flash",
                contents=[
                    types.Part.from_bytes(data=image_bytes, mime_type="image/jpeg"),
                    prompt,
                ],
            )

            raw_text = response.text.strip()
            m = re.search(r'\{.*\}', raw_text, re.DOTALL)
            if m:
                try:
                    return json.loads(m.group())
                except json.JSONDecodeError:
                    pass
            return {"target_object": target_object, "total_count": 0, "detected_objects": []}
        except Exception as e:
            last_err = e
            log(f"[Gemini Count] {idx+1}번째 API 키 실패 (다음 키 시도): {e}")
            continue

    raise Exception(f"모든 Gemini API 키가 실패했습니다. 마지막 오류: {last_err}")




# ══════════════════════════════════════════════════════════
# RFID 쓰기 (리더기 → 빈 태그)
# ══════════════════════════════════════════════════════════
def write_epc_to_reader(epc: str) -> bool:
    """
    시리얼 포트로 EPC Gen2 Write 커맨드 전송.
    ASCII 형식: WRITE <EPC>\\r\\n
    리더기 펌웨어에 따라 커맨드 포맷이 다를 수 있으므로 응답만 로그에 기록.
    """
    ser: serial.Serial | None = st.session_state.get("serial_obj")
    if not ser or not ser.is_open:
        log("[WRITE] 실패: 시리얼 포트 미연결")
        return False
    try:
        ser.write(f"WRITE {epc}\r\n".encode("ascii"))
        time.sleep(0.3)
        resp = ser.read(ser.in_waiting or 32).decode("ascii", errors="ignore").strip()
        log(f"[WRITE] 응답: {resp or '(없음)'}")
        return True
    except UnicodeEncodeError:
        log("[WRITE] 오류: EPC 문자열에 16진수(ASCII)가 아닌 문자(한글 등)가 포함되어 있습니다.")
        return False
    except Exception as e:
        log(f"[WRITE] 오류: {e}")
        return False


# ══════════════════════════════════════════════════════════
# DB 조작 함수
# ══════════════════════════════════════════════════════════
def db_issue_tag(tag_id: str, item_name: str, lot_number: str, quantity: int,
                 category: str = "미분류", mat_number: str = "", batch_code: str = "", location: str = "") -> None:
    """태그 발행 — 이미 존재하면 덮어씀"""
    with get_db() as conn:
        conn.execute(
            "INSERT OR REPLACE INTO tags(tag_id,item_name,lot_number,quantity,issued_at,location,category,mat_number,batch_code) VALUES (?,?,?,?,?,?,?,?,?)",
            (tag_id, item_name, lot_number, quantity, now_str(), location, category, mat_number, batch_code),
        )


def db_get_tag(tag_id: str) -> dict | None:
    """태그 ID로 단일 태그 조회"""
    with get_db() as conn:
        row = conn.execute(
            "SELECT * FROM tags WHERE tag_id=?", (tag_id,)
        ).fetchone()
        return dict(row) if row else None


def db_all_tags() -> list[dict]:
    """전체 태그 목록 (발행일 역순)"""
    with get_db() as conn:
        return [dict(r) for r in conn.execute(
            "SELECT * FROM tags ORDER BY issued_at DESC"
        ).fetchall()]


def db_update_tag_fields(tag_id: str, item_name: str, lot_number: str, quantity: int) -> None:
    """OCR 결과로 태그 정보 갱신"""
    with get_db() as conn:
        conn.execute(
            "UPDATE tags SET item_name=?, lot_number=?, quantity=? WHERE tag_id=?",
            (item_name, lot_number, quantity, tag_id),
        )



def db_update_quantity(tag_id: str, quantity: int) -> None:
    """수량만 갱신"""
    with get_db() as conn:
        conn.execute(
            "UPDATE tags SET quantity=? WHERE tag_id=?", (quantity, tag_id)
        )


def db_log_scan(tag_id: str, matched: bool) -> None:
    """스캔 이력 기록"""
    with get_db() as conn:
        conn.execute(
            "INSERT INTO scan_log(tag_id, scanned_at, matched) VALUES(?,?,?)",
            (tag_id, now_str(), int(matched)),
        )


def db_outgoing(tag_id: str, qty_out: int, reason: str = "소진") -> tuple[bool, str]:
    """원발 차감 + 출고 이력 기록"""
    tag = db_get_tag(tag_id)
    if tag is None:
        return False, "등록되지 않은 태그입니다."
    qty_before = tag["quantity"]
    if qty_out <= 0:
        return False, "출고 수량은 1 이상이어야 합니다."
    if qty_out > qty_before:
        return False, f"출고 수량({qty_out})이 현재 재고({qty_before})를 초과합니다."
    qty_after = qty_before - qty_out
    with get_db() as conn:
        conn.execute("UPDATE tags SET quantity=? WHERE tag_id=?", (qty_after, tag_id))
        conn.execute(
            "INSERT INTO outgoing_log(tag_id,item_name,category,mat_number,qty_out,qty_before,qty_after,out_at,reason) VALUES(?,?,?,?,?,?,?,?,?)",
            (tag_id, tag["item_name"], tag.get("category", ""), tag.get("mat_number", ""),
             qty_out, qty_before, qty_after, now_str(), reason),
        )
    return True, f"출고 완료: {qty_before} → {qty_after} EA"


def db_all_outgoing() -> list[dict]:
    """출고 이력 전체 조회 (최신순)"""
    with get_db() as conn:
        return [dict(r) for r in conn.execute(
            "SELECT * FROM outgoing_log ORDER BY out_at DESC"
        ).fetchall()]


def db_incoming(tag_id: str, qty_in: int, remark: str = "입고") -> tuple[bool, str]:
    """수량 가산 + 입고 이력 기록"""
    tag = db_get_tag(tag_id)
    if tag is None:
        return False, "등록되지 않은 태그입니다."
    qty_before = tag["quantity"]
    if qty_in <= 0:
        return False, "입고 수량은 1 이상이어야 합니다."
    qty_after = qty_before + qty_in
    with get_db() as conn:
        conn.execute("UPDATE tags SET quantity=? WHERE tag_id=?", (qty_after, tag_id))
        conn.execute(
            "INSERT INTO incoming_log(tag_id,item_name,category,mat_number,qty_in,qty_before,qty_after,in_at,remark) VALUES(?,?,?,?,?,?,?,?,?)",
            (tag_id, tag["item_name"], tag.get("category", ""), tag.get("mat_number", ""),
             qty_in, qty_before, qty_after, now_str(), remark),
        )
    return True, f"입고 완료: {qty_before} → {qty_after} EA"


def db_all_incoming() -> list[dict]:
    """입고 이력 전체 조회 (최신순)"""
    with get_db() as conn:
        return [dict(r) for r in conn.execute(
            "SELECT * FROM incoming_log ORDER BY in_at DESC"
        ).fetchall()]


def gen_qr_image(epc: str) -> bytes | None:
    """EPC를 담은 QR코드 PNG 이미지 생성"""
    if not _QR_AVAILABLE:
        return None
    qr = qrcode.QRCode(version=1, box_size=8, border=3,
                       error_correction=qrcode.constants.ERROR_CORRECT_M)
    qr.add_data(epc)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def decode_qr_from_bytes(img_bytes: bytes) -> str | None:
    """카메라 이미지에서 QR코드 디코딩"""
    nparr = np.frombuffer(img_bytes, np.uint8)
    img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    detector = cv2.QRCodeDetector()
    data, _, _ = detector.detectAndDecode(img)
    return data.strip() if data else None


def gen_mock_label_image(cat, mat, name, batch, loc, qty, epc) -> bytes:
    """선택한 태그 정보를 기반으로 가상의 라벨 이미지(JPEG) 생성"""
    img = _PIL_Image.new("RGB", (600, 360), color="#FFFFFF")
    draw = ImageDraw.Draw(img)
    
    # Draw border
    draw.rectangle([10, 10, 590, 350], outline="#1B365D", width=3)
    
    # Try loading a Korean font
    font_title = None
    font_body = None
    font_paths = [
        "C:/Windows/Fonts/malgun.ttf",                     # Windows
        "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",  # Linux (Streamlit Cloud default Korean)
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"   # Linux fallback
    ]
    for p in font_paths:
        if os.path.exists(p):
            try:
                font_title = ImageFont.truetype(p, 20)
                font_body = ImageFont.truetype(p, 15)
                break
            except Exception:
                pass
    if font_title is None:
        try:
            font_title = ImageFont.load_default()
            font_body = ImageFont.load_default()
        except Exception:
            font_title = None
            font_body = None
        
    # Draw text
    draw.text((30, 30), "📦 RFID 자재 태그 라벨 (OCR 테스트용)", fill="#1B365D", font=font_title)
    draw.line([30, 60, 570, 60], fill="#2E5FA3", width=2)
    
    y = 80
    lines = [
        f"자재 분류  {cat}",
        f"자재번호   {mat or '-'}",
        f"자재명     {name}",
        f"배치코드   {batch or '-'}",
        f"저장위치   {loc or '-'}",
        f"수량       {qty} EA",
        f"발행일시   {now_str()}"
    ]
    for line in lines:
        draw.text((30, y), line, fill="#222222", font=font_body)
        y += 28
        
    draw.text((30, y+10), f"EPC: {epc}", fill="#777777", font=font_body)
        
    # Draw a mock barcode or QR placeholder
    draw.rectangle([450, 90, 550, 190], fill="#333333")
    draw.text((455, 200), "출고 스캔용", fill="#333333", font=font_body)
    
    buf = io.BytesIO()
    img.save(buf, format="JPEG")
    return buf.getvalue()


# ══════════════════════════════════════════════════════════
# 스캔 처리
# ══════════════════════════════════════════════════════════
def process_scan(epc: str, source: str = "리더기") -> None:
    """
    수신된 EPC를 세션에 추가.
    동일 세션 내 중복 스캔은 무시하고, DB 매칭 여부를 함께 기록.
    """
    # 이번 세션에서 이미 스캔된 태그는 건너뜀
    if epc in [s["tag_id"] for s in st.session_state["scan_session"]]:
        return

    tag     = db_get_tag(epc)
    matched = tag is not None
    db_log_scan(epc, matched)

    st.session_state["scan_session"].append({
        "tag_id":     epc,
        "category":   tag["category"]   if matched else "-",
        "mat_number": tag.get("mat_number", "") if matched else "-",
        "item_name":  tag["item_name"]  if matched else "❓ 미등록",
        "batch_code": tag.get("batch_code", "") if matched else "-",
        "lot_number": tag["lot_number"] if matched else "-",
        "quantity":   tag["quantity"]   if matched else 0,
        "issued_at":  tag["issued_at"]  if matched else "-",
        "scanned_at": now_str(),
        "matched":    "✅ 매칭" if matched else "⚠️ 미등록",
        "source":     source,
    })
    st.session_state["last_rfid"] = epc
    log(f"[SCAN] {epc} → {'매칭' if matched else '미등록'} ({source})")


# ══════════════════════════════════════════════════════════
# 시리얼 / TCP 백그라운드 리더 스레드
# ══════════════════════════════════════════════════════════
def _parse_epc_from_bytes(raw: bytes) -> str | None:
    """
    리더기 원시 바이트에서 EPC 추출.
    지원 포맷: ASCII 'EPC:xxxx' / 순수 HEX 24자 / 바이너리 패킷
    """
    try:
        text = raw.decode("ascii", errors="ignore").strip()
    except Exception:
        text = ""

    if "EPC:" in text:
        return text.split("EPC:")[-1].strip()[:24]

    clean = re.sub(r'[\r\n\s]', '', text)
    if len(clean) >= 24 and all(c in "0123456789ABCDEFabcdef" for c in clean[:24]):
        return clean[:24].upper()

    # 바이너리 패킷: 헤더 4바이트 이후 12바이트를 EPC로 간주
    if len(raw) >= 16:
        return raw[4:16].hex().upper()

    return None


def _read_loop_serial(ser: serial.Serial) -> None:
    """시리얼 포트 상시 읽기 스레드 — stop_flag 세트 시 종료"""
    log("[SERIAL] 읽기 시작")
    buf = b""
    while not st.session_state["stop_flag"]:
        try:
            if ser.in_waiting:
                buf += ser.read(ser.in_waiting)
                while b"\n" in buf or len(buf) >= 24:
                    if b"\n" in buf:
                        line, buf = buf.split(b"\n", 1)
                    else:
                        line, buf = buf[:24], buf[24:]
                    epc = _parse_epc_from_bytes(line)
                    if epc:
                        process_scan(epc, "시리얼")
            time.sleep(0.05)
        except Exception as e:
            log(f"[SERIAL] 오류: {e}")
            break
    log("[SERIAL] 읽기 종료")


def _read_loop_tcp(sock: socket.socket) -> None:
    """TCP 소켓 상시 읽기 스레드 — stop_flag 세트 시 종료"""
    log("[TCP] 읽기 시작")
    buf = b""
    while not st.session_state["stop_flag"]:
        try:
            sock.settimeout(0.1)
            chunk = sock.recv(256)
            if not chunk:
                break
            buf += chunk
            while b"\n" in buf or len(buf) >= 24:
                if b"\n" in buf:
                    line, buf = buf.split(b"\n", 1)
                else:
                    line, buf = buf[:24], buf[24:]
                epc = _parse_epc_from_bytes(line)
                if epc:
                    process_scan(epc, "TCP")
        except socket.timeout:
            continue
        except Exception as e:
            log(f"[TCP] 오류: {e}")
            break
    log("[TCP] 읽기 종료")


def connect_serial(port: str, baudrate: int) -> bool:
    try:
        ser = serial.Serial(port, baudrate=baudrate, timeout=1)
        st.session_state.update({"serial_obj": ser, "stop_flag": False, "connected": True})
        threading.Thread(target=_read_loop_serial, args=(ser,), daemon=True).start()
        log(f"[CONNECT] 시리얼 {port} @ {baudrate}bps")
        return True
    except Exception as e:
        log(f"[CONNECT] 시리얼 실패: {e}")
        return False


def connect_tcp(host: str, port: int) -> bool:
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.connect((host, port))
        st.session_state.update({"tcp_socket": sock, "stop_flag": False, "connected": True})
        threading.Thread(target=_read_loop_tcp, args=(sock,), daemon=True).start()
        log(f"[CONNECT] TCP {host}:{port}")
        return True
    except Exception as e:
        log(f"[CONNECT] TCP 실패: {e}")
        return False


def disconnect() -> None:
    """연결 해제 및 백그라운드 스레드 정지"""
    st.session_state["stop_flag"] = True
    for key in ("serial_obj", "tcp_socket"):
        obj = st.session_state.get(key)
        if obj:
            try:
                obj.close()
            except Exception:
                pass
            st.session_state[key] = None
    st.session_state["connected"] = False
    log("[CONNECT] 연결 해제")


# ══════════════════════════════════════════════════════════
# YOLO 모델 로더 (폐쇄망 대비)
# ══════════════════════════════════════════════════════════
@st.cache_resource
def load_yolo_model(model_name: str):
    """
    로컬 ./models/ 및 프로젝트 루트 폴더 우선 탐색 → 없으면 ultralytics 자동 다운로드.
    폐쇄망 환경에서는 사전에 ./models/ 폴더에 .pt 파일을 복사해 두면 됨.
    """
    from ultralytics import YOLO

    local_path = os.path.join(MODELS_DIR, model_name)
    if os.path.exists(local_path):
        log(f"[YOLO] 로컬 모델 로드: {local_path}")
        return YOLO(local_path)

    root_path = os.path.join(BASE_DIR, model_name)
    if os.path.exists(root_path):
        log(f"[YOLO] 루트 모델 로드: {root_path}")
        return YOLO(root_path)

    log(f"[YOLO] 로컬 모델 없음 → 자동 다운로드 시도: {model_name}")
    os.makedirs(MODELS_DIR, exist_ok=True)
    model = YOLO(model_name)   # ultralytics 캐시 디렉터리에 다운로드
    return model


# ══════════════════════════════════════════════════════════
# EasyOCR 로더 (캐시로 1회만 로딩)
# ══════════════════════════════════════════════════════════
@st.cache_resource
def load_easyocr_reader():
    """EasyOCR 한국어+영어 리더 초기화 (첫 실행 시 약 20~30초 소요)"""
    import easyocr
    return easyocr.Reader(["ko", "en"], gpu=False)


# ══════════════════════════════════════════════════════════
# 엑셀 생성 (대기업 보고서 수준 서식)
# ══════════════════════════════════════════════════════════
def generate_excel(rows: list[dict]) -> bytes:
    """
    재고조사 결과를 정형화된 xlsx 파일로 변환.
    - 헤더: 네이비 그라디언트 배경 + 흰색 볼드
    - 데이터: 매칭 여부에 따라 초록/주황 배경
    - 열 너비 자동 최적화
    - 요약 행(합계) 추가
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "재고조사결과"
    ws.sheet_view.showGridLines = True

    # ── 색상 팔레트 ────────────────────────────
    C_HDR_DARK  = "1B365D"   # 헤더 좌측 (짙은 네이비)
    C_HDR_LIGHT = "2E5FA3"   # 헤더 우측 (연한 네이비)
    C_OK        = "EBF5EB"   # 매칭 행 배경 (연초록)
    C_ERR       = "FFF3E0"   # 미등록 행 배경 (연주황)
    C_SUM       = "D6E4F0"   # 합계 행 배경 (연파랑)
    C_SUM_FONT  = "1B365D"   # 합계 행 글씨

    # ── 폰트 ───────────────────────────────────
    font_hdr  = Font(name="Malgun Gothic", size=10, bold=True,  color="FFFFFF")
    font_body = Font(name="Malgun Gothic", size=10, bold=False, color="212121")
    font_sum  = Font(name="Malgun Gothic", size=10, bold=True,  color=C_SUM_FONT)

    # ── 채우기 ──────────────────────────────────
    fill_hdr = GradientFill(type="linear", degree=0, stop=(C_HDR_DARK, C_HDR_LIGHT))
    fill_ok  = PatternFill("solid", fgColor=C_OK)
    fill_err = PatternFill("solid", fgColor=C_ERR)
    fill_sum = PatternFill("solid", fgColor=C_SUM)

    # ── 테두리 ──────────────────────────────────
    thin = Side(style="thin",   color="BDBDBD")
    bold = Side(style="medium", color="1B365D")
    bdr_hdr  = Border(left=bold, right=bold, top=bold,  bottom=bold)
    bdr_body = Border(left=thin, right=thin, top=thin,  bottom=thin)
    bdr_sum  = Border(left=bold, right=bold, top=bold,  bottom=bold)

    # ── 정렬 ────────────────────────────────────
    align_c = Alignment(horizontal="center", vertical="center", wrap_text=False)
    align_l = Alignment(horizontal="left",   vertical="center", wrap_text=False)

    # ── 헤더 행 ────────────────────────────────
    headers = ["No.", "자재 분류", "자재번호", "자재명", "배치코드", "수량(EA)", "입고일시", "스캔일시", "매칭 상태"]
    CENTER_COLS = {1, 2, 3, 5, 6, 7, 8, 9}  # 가운데 정렬 열 번호

    ws.append(headers)
    ws.row_dimensions[1].height = 28
    for c_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c_idx)
        cell.font      = font_hdr
        cell.fill      = fill_hdr
        cell.border    = bdr_hdr
        cell.alignment = align_c

    # ── 데이터 행 ──────────────────────────────────────────────────
    for row_idx, row in enumerate(rows, 1):
        vals = [
            row_idx,
            row.get("category", "-"),
            row.get("mat_number", "-"),
            row["item_name"],
            row.get("batch_code", row.get("lot_number", "-")),
            row["quantity"],
            row["issued_at"],
            row["scanned_at"],
            row["matched"],
        ]
        ws.append(vals)
        r = row_idx + 1
        ws.row_dimensions[r].height = 20
        is_ok = row["matched"] == "✅ 매칭"
        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c_idx)
            cell.font      = font_body
            cell.fill      = fill_ok if is_ok else fill_err
            cell.border    = bdr_body
            cell.alignment = align_c if c_idx in CENTER_COLS else align_l

    # ── 합계 행 ────────────────────────────────
    sum_row = len(rows) + 2
    total_qty = sum(r["quantity"] for r in rows)
    matched_cnt = sum(1 for r in rows if r["matched"] == "✅ 매칭")
    ws.append(["합  계", "", "", "", "", total_qty,
               f"매칭: {matched_cnt}건", f"미등록: {len(rows)-matched_cnt}건", ""])
    ws.row_dimensions[sum_row].height = 22
    for c_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=sum_row, column=c_idx)
        cell.font      = font_sum
        cell.fill      = fill_sum
        cell.border    = bdr_sum
        cell.alignment = align_c if c_idx in CENTER_COLS else align_l

    # ── 열 너비 자동 최적화 ─────────────────────
    col_min_widths = {1: 6, 2: 14, 3: 26, 4: 22, 5: 16, 6: 10, 7: 18, 8: 18, 9: 12}
    for col in ws.columns:
        c_idx = col[0].column
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(c_idx)].width = max(
            max_len + 3, col_min_widths.get(c_idx, 12)
        )

    # ── 틀 고정 (헤더 고정) ─────────────────────
    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ══════════════════════════════════════════════════════════
# 사이드바 — 연결 설정 및 Gemini API 키
# ══════════════════════════════════════════════════════════
with st.sidebar:
    st.header("⚙️ 시스템 설정")

    # 클라우드 배포 제약 사항 안내 경고창
    st.info(
        "📡 **RFID 리더기 연결 안내**\\n\\n"
        "대부분의 기능(태그 발행, 모의 스캔, OCR, YOLO, 엑셀 출력 등)은 클라우드에서도 정상 작동합니다.\\n\\n"
        "단, **USB/시리얼(COM 포트) 방식의 리더기**는 로컬 PC에서만 연결 가능합니다. "
        "클라우드에서는 **TCP/IP 방식** 또는 **모의 스캔(시뮬레이션)**을 사용하세요."
    )

    # Gemini API 키 입력
    st.subheader("🤖 Gemini OCR")
    
    # st.secrets 또는 환경변수에서 API 키 자동 로드
    default_key = ""
    toml_keys = []
    if "GEMINI_API_KEYS" in st.secrets:
        tk_list = st.secrets["GEMINI_API_KEYS"]
        if isinstance(tk_list, list):
            toml_keys.extend(tk_list)
        elif isinstance(tk_list, str):
            toml_keys.append(tk_list)
    if "GEMINI_API_KEY" in st.secrets:
        toml_keys.append(st.secrets["GEMINI_API_KEY"])
        
    valid_toml_keys = [tk for tk in toml_keys if tk and "YOUR_GEMINI" not in tk]
    if valid_toml_keys:
        default_key = ", ".join(valid_toml_keys)
    elif os.environ.get("GEMINI_API_KEY"):
        val = os.environ.get("GEMINI_API_KEY")
        if val and "YOUR_GEMINI_API_KEY_HERE" not in val:
            default_key = val

    gemini_api_key_input = st.text_input(
        "API Key",
        type="password",
        value=default_key,
        placeholder="여러 개인 경우 쉼표(,)나 공백으로 구분해 입력",
        help="Google AI Studio에서 무료 발급받은 키를 여러 개 입력 시 한 키가 한도 초과되면 다음 키로 자동 전환됩니다.",
    )
    gemini_api_keys = get_gemini_keys(gemini_api_key_input)
    
    if gemini_api_keys:
        st.success(f"API 키 ✔ ({len(gemini_api_keys)}개 활성화됨)")
    else:
        st.caption("키 없음 — Gemini 컬럼 비활성 ( secrets.toml 설정 가능 )")

    st.divider()

    # RFID 리더기 연결 설정
    st.subheader("📡 RFID 리더기")
    conn_type = st.radio("통신 방식", ["시리얼 (COM)", "TCP/IP"], key="conn_type")

    if conn_type == "시리얼 (COM)":
        ports    = [p.device for p in serial.tools.list_ports.comports()] or ["(포트 없음)"]
        port_sel = st.selectbox("COM 포트", ports)
        baud_sel = st.selectbox("Baudrate", [9600, 19200, 38400, 57600, 115200], index=4)
    else:
        tcp_host = st.text_input("IP 주소", "192.168.1.100")
        tcp_port = st.number_input("포트", value=6000, min_value=1, max_value=65535)

    st.divider()

    # 연결 / 해제 버튼
    if not st.session_state["connected"]:
        if st.button("🔌 리더기 연결", use_container_width=True, type="primary"):
            ok = (connect_serial(port_sel, baud_sel)
                  if conn_type == "시리얼 (COM)"
                  else connect_tcp(tcp_host, int(tcp_port)))
            if ok:
                st.rerun()
            else:
                st.error("연결 실패 — 하단 로그 확인")
    else:
        st.success("✅ 리더기 연결됨")
        if st.session_state["last_rfid"]:
            st.caption(f"최근 태그  \n`{st.session_state['last_rfid']}`")
        if st.button("⛔ 연결 해제", use_container_width=True):
            disconnect()
            st.rerun()

    st.divider()
    if st.button("🧹 더미 데이터 초기화", use_container_width=True, help="기존의 모든 태그 및 이력을 삭제하고 신규 더미 데이터로 리셋합니다."):
        db_reset_dummy()
        st.success("더미 데이터 초기화 완료!")
        st.rerun()


# ══════════════════════════════════════════════════════════
# 탭 구성
# ══════════════════════════════════════════════════════════
tab1, tab_out, tab2, tab3, tab4, tab5 = st.tabs([
    "🏷️ 1. 태그 발행",
    "🔄 2. 입출고 스캔",
    "📡 3. 재고조사 스캔",
    "📷 4. OCR 수량 보완",
    "📊 5. 재고 현황 & 엑셀",
    "🤖 6. YOLO 물체 감지",
])


# ══════════════════════════════════════════════════════════
# 탭 1 — 입고 & 태그 발행
# ══════════════════════════════════════════════════════════
with tab1:
    st.markdown("#### 새 자재 입고 및 RFID 태그 발행")

    _desc_col, _auto_col = st.columns([3, 1])
    with _desc_col:
        st.info("① 품명·Lot·수량 입력  →  ② USB 리더기 위에 빈 태그 올려둠  →  ③ 태그 발행 클릭")
    with _auto_col:
        if st.button("💡 테스트 예제 입력", use_container_width=True, help="테스트용 예제 데이터를 자동으로 입력 필드에 채워줍니다."):
            import random as _rnd
            _sample = _rnd.choice(_DUMMY_TAGS)
            st.session_state["fill_category"] = _sample[5]
            st.session_state["fill_mat_number"] = _sample[2]
            st.session_state["fill_item_name"] = _sample[1]
            st.session_state["fill_batch_code"] = _sample[3] if _sample[3] else f"LOT-{datetime.datetime.now().strftime('%Y%m%d')}"
            st.session_state["fill_location"] = _sample[6]
            st.session_state["fill_quantity"] = int(_sample[4])
            st.rerun()

    with st.form("form_issue"):
        # 1행: 자재 분류 / 자재번호 / 자재명
        r1c1, r1c2, r1c3 = st.columns([1.2, 1.2, 1.5])
        try:
            cat_idx = CATEGORY_LIST.index(st.session_state.get("fill_category", "기자재(초지)"))
        except ValueError:
            cat_idx = 0
        category   = r1c1.selectbox("자재 분류 *", CATEGORY_LIST, index=cat_idx)
        mat_number = r1c2.text_input("자재번호", value=st.session_state.get("fill_mat_number", ""), placeholder="예: MAT-001")
        item_name  = r1c3.text_input("자재명 *", value=st.session_state.get("fill_item_name", ""), placeholder="예: SUS304 판재 2T")

        # 2행: 배치코드 / 저장위치 / 수량
        r2c1, r2c2, r2c3 = st.columns([1.5, 1.5, 1])
        batch_code = r2c1.text_input("배치코드", value=st.session_state.get("fill_batch_code", ""), placeholder="예: LOT-2024-001")
        location   = r2c2.text_input("저장위치", value=st.session_state.get("fill_location", ""), placeholder="예: A구역-1단-03번")
        quantity   = r2c3.number_input("수량 (EA)", min_value=0, value=st.session_state.get("fill_quantity", 0))

        epc_mode   = st.radio("EPC 생성 방식", ["자동 생성 (UUID)", "직접 입력"], horizontal=True)
        manual_epc = ""
        if epc_mode == "직접 입력":
            manual_epc = st.text_input("EPC (24자리 HEX)", max_chars=24)

        submitted = st.form_submit_button("📌 태그 발행", use_container_width=True, type="primary")

    if submitted:
        if not item_name.strip():
            st.error("자재명은 필수입니다.")
        else:
            is_valid = True
            if epc_mode == "직접 입력":
                epc = manual_epc.upper().strip()
                # 16진수 24자리 문자 형태인지 검증
                if not re.match(r'^[0-9A-F]{24}$', epc):
                    st.error("❌ 오류: 직접 입력한 EPC는 반드시 **24자리의 16진수(0-9, A-F)**여야 합니다. (한글, 공백, 특수문자 입력 불가)")
                    is_valid = False
            else:
                epc = gen_epc()

            if is_valid:
                # 리더기 연결 시 실제 태그에 EPC 쓰기 시도
                write_ok = False
                if st.session_state["connected"] and conn_type == "시리얼 (COM)":
                    write_ok = write_epc_to_reader(epc)

                db_issue_tag(epc, item_name.strip(), batch_code.strip(), quantity, category, mat_number.strip(), batch_code.strip(), location.strip())
                log(f"[ISSUE] {epc} / [{category}] {mat_number} {item_name} / {batch_code} / {location} / {quantity}EA")

                st.success("태그 발행 완료!")
                # ── 인쇄 가능한 라벨 렌더링 ──
                _qr_bytes = gen_qr_image(epc)
                _qr_b64 = base64.b64encode(_qr_bytes).decode() if _qr_bytes else ""
                _issued_str = now_str()
                _label_html = f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Malgun Gothic', 'NanumGothic', sans-serif;
          padding: 12px; background: #f8f9fa; }}
  .label {{
    display: flex; gap: 14px; align-items: flex-start;
    border: 2.5px solid #1B365D; border-radius: 10px;
    padding: 14px 16px; background: #fff;
    max-width: 420px; margin: 0 auto;
    box-shadow: 0 2px 8px rgba(0,0,0,0.10);
  }}
  .info {{ flex: 1; }}
  .info .title {{
    font-size: 12px; font-weight: bold; color: #1B365D;
    border-bottom: 1.5px solid #2E5FA3; padding-bottom: 5px; margin-bottom: 8px;
    letter-spacing: 0.5px;
  }}
  .info table {{ width: 100%; border-collapse: collapse; }}
  .info td {{ padding: 3px 4px; font-size: 11px; vertical-align: top; }}
  .info td.lbl {{ color: #555; font-weight: bold; width: 62px; white-space: nowrap; }}
  .info td.val {{ color: #111; word-break: break-all; }}
  .info .epc {{
    margin-top: 8px; font-size: 9px; color: #888;
    word-break: break-all; line-height: 1.3;
    border-top: 1px dashed #ccc; padding-top: 5px;
  }}
  .qr-box {{ text-align: center; flex-shrink: 0; }}
  .qr-box img {{ width: 110px; height: 110px; display: block; }}
  .qr-box .qr-cap {{ font-size: 9px; color: #888; margin-top: 3px; }}
  .print-btn {{
    display: block; width: 420px; max-width: 100%;
    margin: 12px auto 0; padding: 9px 0;
    background: #1B365D; color: #fff; border: none;
    border-radius: 7px; cursor: pointer;
    font-size: 14px; font-family: inherit; font-weight: bold;
    letter-spacing: 0.5px; transition: background 0.2s;
  }}
  .print-btn:hover {{ background: #2E5FA3; }}
  @media print {{
    body {{ background: white; padding: 0; }}
    .print-btn {{ display: none !important; }}
    .label {{ box-shadow: none; border: 2px solid #000;
              max-width: 100%; border-radius: 0; }}
  }}
</style>
</head>
<body>
<div class="label">
  <div class="info">
    <div class="title">📦 RFID 자재 태그 라벨</div>
    <table>
      <tr><td class="lbl">자재 분류</td><td class="val">{category}</td></tr>
      <tr><td class="lbl">자재번호</td><td class="val">{mat_number or "-"}</td></tr>
      <tr><td class="lbl">자재명</td><td class="val">{item_name}</td></tr>
      <tr><td class="lbl">배치코드</td><td class="val">{batch_code or "-"}</td></tr>
      <tr><td class="lbl">저장위치</td><td class="val">{location or "-"}</td></tr>
      <tr><td class="lbl">수량</td><td class="val">{quantity} EA</td></tr>
      <tr><td class="lbl">발행일시</td><td class="val">{_issued_str}</td></tr>
    </table>
    <div class="epc">EPC: {epc}</div>
  </div>
  <div class="qr-box">
    {"<img src='data:image/png;base64," + _qr_b64 + "' alt='QR'>" if _qr_b64 else "<div style='width:110px;height:110px;border:1px dashed #ccc;display:flex;align-items:center;justify-content:center;font-size:10px;color:#aaa;'>QR없음</div>"}
    <div class="qr-cap">출고 스캔용</div>
  </div>
</div>
<button class="print-btn" onclick="window.print()">🖨️ 라벨 인쇄</button>
</body>
</html>"""
                st.components.v1.html(_label_html, height=310, scrolling=False)
                # 다운로드 버튼도 유지
                if _qr_bytes:
                    st.download_button(
                        "⬇️ QR 이미지만 저장 (.png)",
                        data=_qr_bytes,
                        file_name=f"QR_{epc[:8]}.png",
                        mime="image/png",
                    )
                if st.session_state["connected"]:
                    if write_ok:
                        st.info("📡 리더기를 통해 태그에 EPC 기록 완료")
                    else:
                        st.warning("리더기 쓰기 응답 없음 — 태그 위치·모델 커맨드를 확인하세요")
                else:
                    st.warning("리더기 미연결 — DB에만 등록됨 (라벨 인쇄 후 부착 가능)")

    st.markdown("---")
    st.markdown("#### 📋 발행된 태그 전체 목록")
    all_tags = db_all_tags()
    if all_tags:
        _cols = [c for c in ["category", "mat_number", "item_name", "batch_code", "quantity", "issued_at", "location"] if c in all_tags[0]]
        df_tags = pd.DataFrame(all_tags)[_cols]
        _col_names = {
            "category": "자재 분류", "mat_number": "자재번호",
            "item_name": "자재명", "batch_code": "배치코드",
            "quantity": "수량", "issued_at": "발행일시", "location": "위치"
        }
        df_tags.columns = [_col_names[c] for c in _cols]
        st.dataframe(df_tags, use_container_width=True, hide_index=True)
        st.caption(f"총 {len(all_tags)}건 등록됨")
    else:
        st.info("아직 발행된 태그가 없습니다.")


# ══════════════════════════════════════════════════════════
# 탭 2 — 입출고 스캔 (QR / EPC 직접)
# ══════════════════════════════════════════════════════════
OUTGOING_REASONS = ["소진", "파손", "반품", "기타"]
INCOMING_REASONS = ["정상 입고", "반품 회수", "재고 조정", "기타"]

with tab_out:
    st.markdown("#### 🔄 입출고 스캔 — QR코드 또는 EPC 직접 입력")
    st.info("스마트폰 카메라로 자재의 QR코드를 스캔하거나 EPC를 직접 입력해 입고/출고 처리합니다.")

    # ── 입고/출고 모드 선택 ──
    _inout_mode = st.radio(
        "작업 구분", ["📥 입고 (재고 추가)", "📤 출고 (재고 차감)"],
        horizontal=True, key="inout_mode_choice"
    )

    # ── 스캔 입력 방식 선택 ──
    _out_mode = st.radio(
        "스캔 방식", ["📷 QR 카메라 스캔", "⌨️ EPC 직접 입력"],
        horizontal=True, key="out_mode"
    )

    _scanned_epc = ""

    if _out_mode == "📷 QR 카메라 스캔":
        _cam_img = st.camera_input("카메라로 QR코드를 스캔하세요", key="out_cam")
        if _cam_img:
            _decoded = decode_qr_from_bytes(_cam_img.getvalue())
            if _decoded:
                _scanned_epc = _decoded
                st.success(f"✅ QR 인식: `{_scanned_epc}`")
            else:
                st.warning("QR코드를 인식하지 못했습니다. 더 가까이서 촬영해 보세요.")
    else:
        _scanned_epc = st.text_input(
            "EPC 입력", placeholder="예: E004015025A1C00100000001",
            key="out_epc_input"
        ).upper().strip()

    # ── 스캔된 EPC의 자재 정보 표시 ──
    if _scanned_epc:
        _out_tag = db_get_tag(_scanned_epc)
        if _out_tag is None:
            st.error(f"❌ 등록되지 않은 태그입니다: `{_scanned_epc}`")
        else:
            st.markdown("---")
            st.markdown("##### 📦 자재 정보")
            _oc1, _oc2, _oc3, _oc4 = st.columns(4)
            _oc1.metric("자재 분류", _out_tag.get("category", "-"))
            _oc2.metric("자재번호", _out_tag.get("mat_number", "-") or "-")
            _oc3.metric("자재명", _out_tag["item_name"])
            _oc4.metric("현재 재고", f"{_out_tag['quantity']} EA")

            if _inout_mode == "📥 입고 (재고 추가)":
                st.markdown("---")
                _q_col, _r_col = st.columns([1, 1])
                _qty_in = _q_col.number_input(
                    "입고 수량", min_value=1,
                    value=1, key="in_qty"
                )
                _remark = _r_col.selectbox("입고 사유/비고", INCOMING_REASONS, key="in_reason")

                _qty_after_preview = _out_tag["quantity"] + _qty_in
                st.info(
                    f"입고 확인 시: **{_out_tag['quantity']} EA → {_qty_after_preview} EA**"
                )

                if st.button(
                    f"✅ 입고 확인 ({_out_tag['quantity']} → {_qty_after_preview} EA)",
                    type="primary", use_container_width=True, key="in_confirm"
                ):
                    _ok, _msg = db_incoming(_scanned_epc, _qty_in, _remark)
                    if _ok:
                        log(f"[IN] {_scanned_epc} | {_out_tag['item_name']} | +{_qty_in}EA | {_remark}")
                        st.success(f"✅ {_msg}")
                        st.rerun()
                    else:
                        st.error(f"❌ {_msg}")

            else:
                # 📤 출고 모드
                if _out_tag["quantity"] == 0:
                    st.error("⛔ 현재 재고가 0입니다. 출고할 수 없습니다.")
                else:
                    st.markdown("---")
                    _q_col, _r_col = st.columns([1, 1])
                    _qty_out = _q_col.number_input(
                        "출고 수량", min_value=1,
                        max_value=int(_out_tag["quantity"]),
                        value=1, key="out_qty"
                    )
                    _reason = _r_col.selectbox("출고 사유", OUTGOING_REASONS, key="out_reason")

                    _qty_after_preview = _out_tag["quantity"] - _qty_out
                    st.info(
                        f"출고 확인 시: **{_out_tag['quantity']} EA → {_qty_after_preview} EA** "
                        f"({'⚠️ 재고 소진!' if _qty_after_preview == 0 else ''})"
                    )

                    if st.button(
                        f"✅ 출고 확인 ({_out_tag['quantity']} → {_qty_after_preview} EA)",
                        type="primary", use_container_width=True, key="out_confirm"
                    ):
                        _ok, _msg = db_outgoing(_scanned_epc, _qty_out, _reason)
                        if _ok:
                            log(f"[OUT] {_scanned_epc} | {_out_tag['item_name']} | -{_qty_out}EA | {_reason}")
                            st.success(f"✅ {_msg}")
                            if _qty_after_preview == 0:
                                st.warning("⚠️ 해당 자재의 재고가 모두 소진되었습니다!")
                            st.rerun()
                        else:
                            st.error(f"❌ {_msg}")

    st.markdown("---")
    st.markdown("#### 📋 입출고 이력")
    _hist_tab_in, _hist_tab_out = st.tabs(["📥 입고 이력", "📤 출고 이력"])

    with _hist_tab_in:
        _in_logs = db_all_incoming()
        if _in_logs:
            _df_in = pd.DataFrame(_in_logs)[[
                "in_at", "category", "mat_number", "item_name",
                "qty_in", "qty_before", "qty_after", "remark"
            ]]
            _df_in.columns = ["입고일시", "분류", "자재번호", "자재명", "입고량", "입고전", "입고후", "비고"]
            st.dataframe(_df_in, use_container_width=True, hide_index=True)
            st.caption(f"총 {len(_in_logs)}건 입고")
        else:
            st.info("아직 입고 이력이 없습니다.")

    with _hist_tab_out:
        _out_logs = db_all_outgoing()
        if _out_logs:
            _df_out = pd.DataFrame(_out_logs)[[
                "out_at", "category", "mat_number", "item_name",
                "qty_out", "qty_before", "qty_after", "reason"
            ]]
            _df_out.columns = ["출고일시", "분류", "자재번호", "자재명", "출고량", "출고전", "출고후", "사유"]
            st.dataframe(_df_out, use_container_width=True, hide_index=True)
            st.caption(f"총 {len(_out_logs)}건 출고")
        else:
            st.info("아직 출고 이력이 없습니다.")


# ══════════════════════════════════════════════════════════
# 탭 2 — 재고조사 스캔
# ══════════════════════════════════════════════════════════
with tab2:
    st.markdown("#### 재고조사 스캔")
    st.info("카트를 밀며 리더기로 태그를 스캔하면 아래 목록에 실시간으로 쌓입니다.")

    c1, c2 = st.columns([2, 2])
    with c1:
        if st.button("🔴 모의 스캔 (시뮬레이션)", use_container_width=True):
            all_tag_ids  = [t["tag_id"] for t in db_all_tags()]
            scanned_ids  = [s["tag_id"] for s in st.session_state["scan_session"]]
            remaining    = [t for t in all_tag_ids if t not in scanned_ids]

            if all_tag_ids:
                pick = remaining[0] if remaining else all_tag_ids[0]
                process_scan(pick, "시뮬레이션")
                st.toast(f"스캔: {pick[:16]}…", icon="📡")
            else:
                # 등록 태그 없으면 미등록 태그로 시뮬레이션
                fake = gen_epc()
                process_scan(fake, "시뮬레이션(미등록)")
                st.toast("미등록 태그 감지!", icon="⚠️")

    with c2:
        if st.button("🧹 스캔 세션 초기화", use_container_width=True):
            st.session_state["scan_session"]    = []
            st.session_state["scan_count_prev"] = 0
            st.rerun()

    # ── 리더기 연결 중 폴링 ──────────────────────
    # time.sleep + st.rerun 대신: 새 태그가 들어왔을 때만 토스트 알림
    # 화면 자동 갱신은 st_autorefresh 없이 버튼 클릭 때만 실행
    if st.session_state["connected"]:
        current_count = len(st.session_state["scan_session"])
        prev_count    = st.session_state["scan_count_prev"]

        if current_count > prev_count:
            # 새 태그가 추가된 경우에만 토스트 + 카운트 갱신
            new_tags = st.session_state["scan_session"][prev_count:]
            for nt in new_tags:
                icon = "✅" if nt["matched"] == "✅ 매칭" else "⚠️"
                st.toast(f"{icon} {nt['item_name']} 스캔됨", icon=icon)
            st.session_state["scan_count_prev"] = current_count

        st.success(f"📡 리더기 수신 대기 중 — 현재 {current_count}건 스캔됨")

        # 수동 갱신 버튼 (깜빡임 없이 사용자가 원할 때만)
        if st.button("🔄 목록 갱신", help="리더기에서 새 태그를 읽었으면 클릭하세요"):
            st.rerun()

    # ── 스캔 결과 표시 ─────────────────────────
    session = st.session_state["scan_session"]
    if session:
        df_scan = pd.DataFrame(session)[
            ["tag_id", "item_name", "lot_number", "quantity", "issued_at", "scanned_at", "matched"]
        ]
        df_scan.columns = ["태그 ID", "품명", "Lot 번호", "수량", "입고일시", "스캔일시", "매칭 상태"]
        st.dataframe(df_scan, use_container_width=True, hide_index=True)

        total     = len(session)
        matched   = sum(1 for s in session if s["matched"] == "✅ 매칭")
        unmatched = total - matched

        m1, m2, m3 = st.columns(3)
        m1.metric("총 스캔",    f"{total} 건")
        m2.metric("매칭 성공",  f"{matched} 건")
        m3.metric("미등록 태그", f"{unmatched} 건",
                  delta=f"-{unmatched}" if unmatched else None,
                  delta_color="inverse")
    else:
        st.info("스캔 결과가 없습니다. 모의 스캔 버튼 또는 실제 리더기를 사용해보세요.")


# ══════════════════════════════════════════════════════════
# 탭 3 — OCR 엔진 비교 (EasyOCR vs Gemini 2.5 Flash)
# ══════════════════════════════════════════════════════════
with tab3:
    st.markdown("#### 📷 OCR 수량 보완 — EasyOCR vs Gemini")
    st.caption("자재 라벨 사진을 인식하여 품명, Lot 번호, 수량을 파악하고 DB를 업데이트합니다.")

    # 업데이트할 태그 선택
    all_tag_list = db_all_tags()
    tag_ids      = [t["tag_id"] for t in all_tag_list]

    if not tag_ids:
        st.warning("먼저 1번 탭에서 태그를 발행해주세요.")
    else:
        default_idx = (
            tag_ids.index(st.session_state["last_rfid"])
            if st.session_state["last_rfid"] in tag_ids else 0
        )
        
        c_sel, c_eng = st.columns([1, 1])
        with c_sel:
            sel_tag  = st.selectbox("업데이트할 태그 선택", tag_ids, index=default_idx, key="ocr_tag_sel")
            tag_info = db_get_tag(sel_tag)
            if tag_info:
                st.caption(
                    f"현재 DB → 품명: **{tag_info['item_name']}** | "
                    f"Lot: **{tag_info['lot_number']}** | "
                    f"수량: **{tag_info['quantity']} EA**"
                )
        with c_eng:
            ocr_engine = st.radio(
                "사용할 OCR 엔진 선택",
                ["EasyOCR (로컬·무료)", "Gemini 2.5 Flash (AI 분석)"],
                horizontal=True,
                key="ocr_engine_choice"
            )

        st.markdown("---")

        # 이미지 입력 방식 선택
        ocr_mode = st.radio(
            "이미지 입력 방식", ["📷 카메라 촬영", "📁 파일 업로드", "💡 테스트용 샘플 이미지"],
            horizontal=True, key="ocr_input_mode"
        )
        
        img_file = None
        if ocr_mode == "📷 카메라 촬영":
            img_file = st.camera_input("자재 라벨을 촬영해주세요")
        elif ocr_mode == "📁 파일 업로드":
            img_file = st.file_uploader("이미지 파일 업로드", type=["jpg","jpeg","png","bmp","webp"], key="ocr_upload")
        else:
            if st.button("💡 선택한 태그의 샘플 라벨 이미지 생성"):
                st.session_state["ocr_sample_bytes"] = gen_mock_label_image(
                    tag_info.get("category", "미분류"),
                    tag_info.get("mat_number", ""),
                    tag_info["item_name"],
                    tag_info.get("batch_code", ""),
                    tag_info.get("location", ""),
                    tag_info["quantity"],
                    sel_tag
                )
            if "ocr_sample_bytes" in st.session_state:
                img_file = io.BytesIO(st.session_state["ocr_sample_bytes"])
                st.info("샘플 라벨 이미지가 가상으로 생성되었습니다. 아래 분석 결과를 확인해 보세요.")

        if img_file:
            img_bytes = img_file.getvalue()
            
            col_img, col_res = st.columns([1, 1])
            with col_img:
                st.markdown("**입력 이미지**")
                st.image(img_bytes, use_container_width=True)

            with col_res:
                if ocr_engine == "EasyOCR (로컬·무료)":
                    st.markdown("#### 🔵 EasyOCR &nbsp; `로컬·무료`")
                    st.caption("인터넷 불필요 · 첫 실행 시 모델 로딩 약 20~30초")

                    with st.status("EasyOCR 분석 중...", expanded=True) as easy_status:
                        try:
                            st.write("모델 로딩 확인 중...")
                            reader = load_easyocr_reader()
                            st.write("이미지 디코딩 중...")
                            cv2_img    = cv2.imdecode(np.frombuffer(img_bytes, np.uint8), cv2.IMREAD_COLOR)
                            st.write("텍스트 인식 중...")
                            easy_texts = reader.readtext(cv2_img, detail=0)
                            easy_qty   = extract_qty_from_texts(easy_texts)
                            easy_ok    = True
                            log("[EasyOCR] 완료")
                            easy_status.update(label="EasyOCR 완료 ✔", state="complete")
                        except Exception as e:
                            easy_texts, easy_qty, easy_ok = [], None, False
                            easy_status.update(label=f"EasyOCR 오류: {e}", state="error")

                    if easy_ok:
                        st.markdown("**인식된 텍스트**")
                        for t in easy_texts:
                            st.write(f"- {t}")
                        st.markdown(
                            f"**→ 추출 수량: `{easy_qty if easy_qty is not None else '미인식'}`**"
                        )
                        with st.expander("EasyOCR 결과로 저장", expanded=True):
                            ea, eb, ec = st.columns(3)
                            e_name = ea.text_input("품명", value=tag_info["item_name"] if tag_info else "", key="e_name")
                            e_lot  = eb.text_input("Lot",  value=tag_info["lot_number"] if tag_info else "", key="e_lot")
                            e_qty  = ec.number_input("수량", min_value=0, value=easy_qty or 0, key="e_qty")
                            if st.button("💾 EasyOCR 결과 저장", use_container_width=True, key="save_easy"):
                                db_update_tag_fields(sel_tag, e_name, e_lot, e_qty)
                                log(f"[EasyOCR SAVE] {sel_tag} | {e_qty}EA")
                                st.success(f"저장 완료: {e_qty} EA")
                                st.rerun()

                else:  # Gemini 2.5 Flash
                    st.markdown("#### 🟡 Gemini 2.5 Flash &nbsp; `API`")
                    st.caption("품명·Lot·수량 한 번에 파악 · Gemini API 키 필요")

                    if not gemini_api_keys:
                        st.warning("사이드바에 Gemini API 키를 입력하면 활성화됩니다.")
                        st.markdown("[API 키 발급 → Google AI Studio](https://aistudio.google.com/apikey)")
                    else:
                        with st.status("Gemini 2.5 Flash 분석 중...", expanded=True) as g_status:
                            try:
                                st.write("Gemini API 호출 중...")
                                g_result = gemini_ocr(img_bytes, gemini_api_keys)
                                g_ok     = True
                                log(f"[Gemini] 완료: {g_result}")
                                g_status.update(label="Gemini 완료 ✔", state="complete")
                            except Exception as e:
                                g_result, g_ok = {}, False
                                g_status.update(
                                    label=f"Gemini 오류: {type(e).__name__}", state="error"
                                )
                                st.error(
                                    f"**Gemini API 오류**  \n"
                                    f"`{e}`  \n"
                                    "API 키·네트워크 상태를 확인해 주세요."
                                )

                        if g_ok:
                            st.markdown("**인식된 전체 텍스트**")
                            st.code(g_result.get("raw_text") or "(없음)")
                            st.markdown(
                                f"**→ 품명:** `{g_result.get('item_name') or '미인식'}`  \n"
                                f"**→ Lot:** `{g_result.get('lot') or '미인식'}`  \n"
                                f"**→ 수량:** `{g_result.get('quantity') if g_result.get('quantity') is not None else '미인식'}`"
                            )
                            with st.expander("Gemini 결과로 저장", expanded=True):
                                ga, gb, gc = st.columns(3)
                                g_name = ga.text_input(
                                    "품명", key="g_name",
                                    value=g_result.get("item_name") or (tag_info["item_name"] if tag_info else ""),
                                )
                                g_lot = gb.text_input(
                                    "Lot", key="g_lot",
                                    value=g_result.get("lot") or (tag_info["lot_number"] if tag_info else ""),
                                )
                                g_qty = gc.number_input(
                                    "수량", min_value=0, key="g_qty",
                                    value=int(g_result["quantity"]) if g_result.get("quantity") else
                                          (tag_info["quantity"] if tag_info else 0),
                                )
                                if st.button("💾 Gemini 결과 저장", use_container_width=True, key="save_gemini"):
                                    db_update_tag_fields(sel_tag, g_name, g_lot, g_qty)
                                    log(f"[Gemini SAVE] {sel_tag} | {g_name} | {g_lot} | {g_qty}EA")
                                    st.success(f"저장 완료: {g_qty} EA")
                                    st.rerun()


# ══════════════════════════════════════════════════════════
# 탭 4 — 재고 현황 & 엑셀 다운로드
# ══════════════════════════════════════════════════════════
with tab4:
    st.markdown("#### 📊 재고 현황 및 엑셀 다운로드")

    session = st.session_state["scan_session"]
    if not session:
        st.info("2번 탭에서 재고조사 스캔을 먼저 진행해주세요.")
    else:
        df_result = pd.DataFrame(session)[
            ["category", "tag_id", "item_name", "lot_number", "quantity", "issued_at", "scanned_at", "matched"]
        ]
        df_result.columns = ["자재 분류", "태그 ID", "품명", "Lot 번호", "수량", "입고일시", "스캔일시", "매칭 상태"]
        st.dataframe(df_result, use_container_width=True, hide_index=True)

        total_cnt  = len(session)
        match_cnt  = sum(1 for s in session if s["matched"] == "✅ 매칭")
        total_qty  = sum(s["quantity"] for s in session)
        unmatch_cnt = total_cnt - match_cnt

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("총 스캔 건수",  f"{total_cnt} 건")
        m2.metric("매칭 성공",     f"{match_cnt} 건")
        m3.metric("총 수량 합계",  f"{total_qty} EA")
        m4.metric("미등록 태그",   f"{unmatch_cnt} 건",
                  delta=f"-{unmatch_cnt}" if unmatch_cnt else None,
                  delta_color="inverse")

        st.markdown("---")
        st.download_button(
            label="📥 재고조사 결과 엑셀 다운로드 (.xlsx)",
            data=generate_excel(session),
            file_name=f"RFID_재고조사_{datetime.date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )


# ══════════════════════════════════════════════════════════
# 탭 5 — YOLO 물체 감지 (폐쇄망 대비 로컬 모델 우선)
# ══════════════════════════════════════════════════════════
with tab5:
    st.markdown("#### 🤖 물체 감지 & 수량 카운팅 — YOLO vs Gemini")
    st.info(
        "촬영하거나 업로드한 사진을 분석하여 물품의 수량을 카운팅합니다.  \n"
        "**YOLO(로컬)** 또는 **Gemini 2.5 Flash(AI 분석)** 엔진 중 선택할 수 있습니다."
    )

    # ── 엔진 및 옵션 선택 ────────────────────────
    engine_choice = st.radio(
        "감지 엔진 선택",
        ["YOLO (로컬 모델)", "Gemini 2.5 Flash (AI 분석)"],
        horizontal=True,
        key="yolo_engine_choice"
    )

    op1, op2, op3 = st.columns(3)

    if engine_choice == "YOLO (로컬 모델)":
        model_name = op1.selectbox(
            "YOLO 모델",
            YOLO_CANDIDATES,
            index=0,
            help="로컬 ./models/ 우선 탐색 → 없으면 자동 다운로드",
        )
        conf_thresh = op2.slider("신뢰도 임계값", 0.10, 0.90, 0.40, 0.05,
                                  help="낮을수록 더 많이 감지, 높을수록 정확")
        target_cls = op3.text_input(
            "특정 클래스만 카운팅 (선택)",
            placeholder="예: bottle, cup",
            help="빈칸 = 전체 합산. 쉼표로 여러 개 입력",
        )

        # 로컬 모델 파일 존재 여부 안내
        local_pt = os.path.join(MODELS_DIR, model_name)
        if os.path.exists(local_pt):
            st.caption(f"✅ 로컬 모델 감지됨: `{local_pt}`")
        else:
            st.caption(f"⚠️ 로컬 모델 없음 — 첫 실행 시 자동 다운로드됩니다 (`{local_pt}`에 복사하면 오프라인 사용 가능)")
    
    else:  # Gemini 2.5 Flash
        target_cls = op1.text_input(
            "감지 및 카운팅할 물체 입력 *",
            value="상자",
            help="예: 상자, 제품, 컵, 볼트 등"
        )
        if gemini_api_keys:
            op2.success(f"API 키 확인됨 ✔ ({len(gemini_api_keys)}개)")
        else:
            op2.error("사이드바에서 API 키를 입력해 주세요.")
        op3.caption("Gemini는 사전 학습 없이 이미지 속 물품을 실시간으로 감지하고 상자를 그려줍니다.")

    st.markdown("---")

    # ── 이미지 입력 ─────────────────────────────
    yolo_mode = st.radio(
        "이미지 입력 방식", ["📷 카메라 촬영", "📁 파일 업로드", "💡 테스트용 샘플 이미지"],
        horizontal=True, key="yolo_input_mode"
    )
    
    img_file_yolo = None
    if yolo_mode == "📷 카메라 촬영":
        img_file_yolo = st.camera_input("자재를 촬영해주세요", key="yolo_cam")
    elif yolo_mode == "📁 파일 업로드":
        img_file_yolo = st.file_uploader("이미지 업로드", type=["jpg","jpeg","png","bmp","webp"], key="yolo_upload")
    else:
        sample_path = os.path.join(BASE_DIR, "Roll images.jpg")
        if os.path.exists(sample_path):
            if st.button("💡 Roll images.jpg 샘플 로드"):
                with open(sample_path, "rb") as f:
                    st.session_state["yolo_sample_bytes"] = f.read()
            if "yolo_sample_bytes" in st.session_state:
                img_file_yolo = io.BytesIO(st.session_state["yolo_sample_bytes"])
                st.info("샘플 롤 이미지가 로드되었습니다. 아래 분석 결과를 확인해 보세요.")
        else:
            st.error(f"샘플 롤 이미지 파일(`{sample_path}`)을 찾을 수 없습니다. 프로젝트 루트에 파일이 있는지 확인해 주세요.")

    if img_file_yolo:
        img_bytes_yolo = img_file_yolo.getvalue()
        img_np = cv2.imdecode(np.frombuffer(img_bytes_yolo, np.uint8), cv2.IMREAD_COLOR)

        det_ok = False
        total_det = 0
        annotated_image = None
        filtered_results = {}

        if engine_choice == "YOLO (로컬 모델)":
            with st.status(f"YOLO 감지 중 ({model_name})...", expanded=True) as yolo_status:
                try:
                    st.write("모델 로딩 중...")
                    model = load_yolo_model(model_name)
                    st.write("이미지 추론 중...")
                    results = model(img_np, conf=conf_thresh, verbose=False)
                    result = results[0]
                    det_ok = True
                    
                    # 클래스별 개수 집계
                    names = model.names
                    class_counts = {}
                    for cls_id in result.boxes.cls.tolist():
                        label = names[int(cls_id)]
                        class_counts[label] = class_counts.get(label, 0) + 1
                    
                    # 특정 클래스 필터 적용
                    filter_cls = [c.strip().lower() for c in target_cls.split(",") if c.strip()]
                    filtered_results = {k: v for k, v in class_counts.items()
                                       if not filter_cls or k.lower() in filter_cls}
                    total_det = sum(filtered_results.values())
                    
                    # 시각화 이미지 (BGR -> RGB)
                    annotated_image = cv2.cvtColor(result.plot(), cv2.COLOR_BGR2RGB)
                    
                    log(f"[YOLO] {len(result.boxes)}개 박스 감지 완료 (필터 후 {total_det}개)")
                    yolo_status.update(label=f"YOLO 감지 완료 — 필터 후 {total_det}개", state="complete")
                except Exception as e:
                    det_ok = False
                    yolo_status.update(label=f"YOLO 오류: {e}", state="error")
                    st.error(f"**YOLO 실행 오류**  \n`{e}`")
                    
        else:  # Gemini 2.5 Flash
            if not gemini_api_keys:
                st.error("사이드바에 Gemini API 키를 입력해야 이용 가능합니다.")
            elif not target_cls.strip():
                st.warning("감지할 물체 이름을 입력해 주세요.")
            else:
                with st.status("Gemini 2.5 Flash 분석 중...", expanded=True) as g_status:
                    try:
                        st.write("Gemini API 호출 및 물체 카운팅 중...")
                        g_result = gemini_count_objects(img_bytes_yolo, target_cls.strip(), gemini_api_keys)
                        
                        total_det = g_result.get("total_count", 0)
                        detected_objs = g_result.get("detected_objects", [])
                        filtered_results = {target_cls.strip(): total_det}
                        det_ok = True
                        
                        # OpenCV 이미지 위에 바운딩 박스 그리기
                        h, w, _ = img_np.shape
                        img_draw = img_np.copy()
                        for obj in detected_objs:
                            box = obj.get("box_2d")
                            if box and len(box) == 4:
                                ymin, xmin, ymax, xmax = box
                                # 0~1000 상대 좌표를 픽셀 단위로 변환
                                y1 = int(ymin * h / 1000)
                                x1 = int(xmin * w / 1000)
                                y2 = int(ymax * h / 1000)
                                x2 = int(xmax * w / 1000)
                                
                                # 녹색 사각형 및 텍스트 그리기
                                cv2.rectangle(img_draw, (x1, y1), (x2, y2), (0, 255, 0), 2)
                                label = obj.get("label", target_cls.strip())
                                cv2.putText(img_draw, label, (x1, y1 - 10), 
                                            cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 1)
                        
                        # RGB 변환
                        annotated_image = cv2.cvtColor(img_draw, cv2.COLOR_BGR2RGB)
                        
                        log(f"[Gemini] '{target_cls.strip()}' {total_det}개 감지 완료")
                        g_status.update(label=f"Gemini 감지 완료 — {total_det}개", state="complete")
                    except Exception as e:
                        det_ok = False
                        g_status.update(label=f"Gemini 오류: {e}", state="error")
                        st.error(f"**Gemini API 오류**  \n`{e}`")

        if det_ok:
            # ── 결과 표시 ────────────────────────
            col_img, col_stat = st.columns([3, 2])

            with col_img:
                st.markdown("**감지 결과 이미지**")
                if annotated_image is not None:
                    st.image(annotated_image, use_container_width=True)
                else:
                    st.warning("표시할 이미지가 없습니다.")

            with col_stat:
                st.markdown("**종류별 감지 결과**")
                if total_det > 0:
                    df_det = pd.DataFrame(
                        filtered_results.items(), columns=["클래스(감지 물체)", "개수"]
                    ).sort_values("개수", ascending=False)
                    st.dataframe(df_det, use_container_width=True, hide_index=True)
                    st.metric("총 감지 개수", f"{total_det} 개")
                else:
                    st.warning("감지된 물체가 없습니다. 신뢰도를 낮추거나 클래스 필터명을 확인해 주세요.")

            # ── 수량 태그에 저장 ─────────────────
            if total_det > 0:
                st.markdown("---")
                st.markdown("#### 감지 수량을 태그에 저장")

                tag_ids_yolo = [t["tag_id"] for t in db_all_tags()]
                if not tag_ids_yolo:
                    st.warning("저장할 태그가 없습니다. 1번 탭에서 태그를 먼저 발행하세요.")
                else:
                    default_yolo_idx = (
                        tag_ids_yolo.index(st.session_state["last_rfid"])
                        if st.session_state["last_rfid"] in tag_ids_yolo else 0
                    )
                    y1, y2, y3 = st.columns([3, 2, 2])
                    sel_tag_yolo = y1.selectbox(
                        "저장할 태그", tag_ids_yolo,
                        index=default_yolo_idx, key="yolo_tag"
                    )
                    yolo_qty = y2.number_input(
                        "저장할 수량", min_value=0, value=int(total_det), key="yolo_qty"
                    )
                    y3.markdown("<br>", unsafe_allow_html=True)
                    if y3.button("💾 수량 저장", use_container_width=True, key="save_yolo", type="primary"):
                        db_update_quantity(sel_tag_yolo, yolo_qty)
                        log(f"[DETECTION SAVE] {sel_tag_yolo} → {yolo_qty}EA")
                        st.success(f"저장 완료: {yolo_qty} EA → `{sel_tag_yolo}`")
                        st.rerun()

    # ── 하단 학습 가이드 추가 ─────────────────────
    st.markdown("---")
    with st.expander("🎓 YOLO 커스텀 모델 학습 방법 가이드", expanded=False):
        st.markdown(
            """
            #### 1. 데이터 수집 및 라벨링
            * **데이터 준비**: 감지하고자 하는 고유한 부품/상자를 다양한 조명과 각도에서 촬영합니다. (최소 100장 이상 권장)
            * **라벨링**: 무료 온라인 툴인 [Roboflow](https://roboflow.com/)를 권장합니다.
              * 프로젝트 타입을 `Object Detection`으로 설정하고 이미지를 업로드합니다.
              * 마우스로 각 물건에 사각형 상자를 그리고 이름을 입력합니다. (예: `box`)
              * 라벨링 완료 후 `Export` 단계에서 **YOLOv8 / YOLOv11** 포맷을 선택해 데이터셋을 다운로드합니다.

            #### 2. 로컬 학습 실행
            * 프로젝트 폴더에 자동으로 생성된 `train_yolo.py` 스크립트를 활용합니다.
            * 다운로드한 데이터셋을 `dataset/` 폴더 하위에 위치시킨 뒤, 윈도우 터미널(PowerShell)에서 다음 명령어로 학습을 실행합니다:
              ```bash
              python train_yolo.py
              ```
            * 학습이 진행되며 결과는 `runs/detect/train/weights/best.pt` 경로에 최적 가중치 파일로 저장됩니다.

            #### 3. 대시보드 연동
            * 생성된 `best.pt` 파일의 이름을 구분하기 좋게 수정(예: `box_detect_best.pt`)하여 `./models/` 폴더로 복사합니다.
            * 대시보드 화면을 새로고침하면 `YOLO 모델` 선택창에서 새로 복사한 커스텀 모델을 바로 골라 사용할 수 있습니다!
            """
        )



# ══════════════════════════════════════════════════════════
# 하단 시스템 로그
# ══════════════════════════════════════════════════════════
with st.expander("🖥️ 시스템 로그", expanded=False):
    msgs = st.session_state["log_messages"]
    if msgs:
        st.code("\n".join(reversed(msgs)), language=None)
    else:
        st.write("로그 없음")
